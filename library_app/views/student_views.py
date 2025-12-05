from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from ..models import Student, Centre, CustomUser, School
import csv
import openpyxl
from io import TextIOWrapper
import random


def is_authorized(user):
    """Check if user has permission to manage students"""
    return user.is_superuser or user.is_librarian


@login_required
def get_schools_by_centre(request):
    """AJAX endpoint to get schools for a specific centre — ALREADY PERFECT"""
    centre_id = request.GET.get('centre_id')
    if not centre_id:
        return JsonResponse({'schools': []})
    try:
        schools = School.objects.filter(centre_id=centre_id).values('id', 'name')
        return JsonResponse({'schools': list(schools)})
    except Exception as e:
        print(f"Error fetching schools for centre {centre_id}: {str(e)}")
        return JsonResponse({'schools': []}, status=500)



@login_required
def manage_students(request):
    """Main view to list and manage students — UPDATED FOR CHILD_ID LOGIN"""
    if not is_authorized(request.user):
        messages.error(request, "You do not have permission to access this page.")
        print(f"Unauthorized access attempt by user {request.user.id} to manage_students")
        return redirect('dashboard')

    # Get students based on user role
    if request.user.is_site_admin or request.user.is_superuser:
        students = Student.objects.select_related('user', 'centre', 'school').all().order_by('name')
    else:
        students = Student.objects.select_related('user', 'centre', 'school').filter(
            centre=request.user.centre
        ).order_by('name')

    # Search functionality — now includes login ID
    query = request.GET.get('q', '').strip()
    if query:
        students = students.filter(
            Q(name__icontains=query) |
            Q(child_ID__icontains=query) |
            Q(user__login_id__icontains=query) |  # Search by login ID too
            Q(school__name__icontains=query)
        )

    # Pagination
    items_per_page_options = [10, 25, 50, 100]
    items_per_page = request.GET.get('items_per_page', '25')
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in items_per_page_options:
            items_per_page = 25
    except ValueError:
        items_per_page = 25

    paginator = Paginator(students, items_per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get schools and centres for dropdowns
    if request.user.is_superuser or request.user.is_site_admin:
        centres = Centre.objects.all()
    else:
        centres = [request.user.centre] if request.user.centre else []

    schools = School.objects.filter(centre__in=centres) if centres else School.objects.none()

    # Determine permissions
    can_manage = request.user.is_superuser or request.user.is_librarian or request.user.is_site_admin

    context = {
        'students': page_obj,
        'query': query,
        'items_per_page': items_per_page,
        'items_per_page_options': items_per_page_options,
        'schools': schools,
        'centres': centres,
        'grade_choices': Student.GRADE_CHOICES,
        'can_manage': can_manage,
    }
    return render(request, 'students/student_list.html', context)

@login_required
def add_student(request):
    """Add a single student — UPDATED FOR NEW MODELS"""
    if not is_authorized(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('manage_students')

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        child_ID = request.POST.get('child_ID', '').strip()
        school_id = request.POST.get('school', '').strip()
        grade = request.POST.get('grade', '').strip()
        centre_id = request.POST.get('centre') if request.user.is_superuser else str(request.user.centre.id)

        # Validation
        if not all([first_name, child_ID, school_id]):
            messages.error(request, "First name, Child ID, and School are required.")
            return redirect('manage_students')
        
        if Student.objects.filter(child_ID=child_ID).exists():
            messages.error(request, f"Child ID {child_ID} already exists.")
            return redirect('manage_students')

        try:
            centre = Centre.objects.get(id=centre_id) if request.user.is_superuser else request.user.centre
            school = get_object_or_404(School, id=school_id, centre=centre)

            with transaction.atomic():
                # Create Student first → triggers user creation via signal
                student = Student(
                    child_ID=child_ID,
                    name=f"{first_name} {last_name}".strip(),
                    centre=centre,
                    school=school,
                    grade=grade if grade in dict(Student.GRADE_CHOICES) else None
                )
                student.save()  # This creates the CustomUser with login_id = child_ID

                # Set initial password = child_ID
                student.user.set_password(child_ID)
                student.user.force_password_change = True
                student.user.save()

                messages.success(
                    request,
                    f"Student {student.name} added! Login ID: {child_ID}, Password: {child_ID} (must change on first login)"
                )
                return redirect('manage_students')

        except Exception as e:
            messages.error(request, f"Error creating student: {str(e)}")
            return redirect('manage_students')
    
    return redirect('manage_students')

@login_required
def bulk_upload_students(request):
    """Bulk upload — UPDATED FOR NEW MODELS"""
    if not is_authorized(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('manage_students')

    if request.method != 'POST' or 'file' not in request.FILES:
        messages.error(request, "Please select a file to upload.")
        return redirect('manage_students')

    school_id = request.POST.get('bulk_school', '').strip()
    centre_id = request.POST.get('bulk_centre', '').strip()

    if not school_id:
        messages.error(request, "Please select a school.")
        return redirect('manage_students')

    if request.user.is_superuser:
        if not centre_id:
            messages.error(request, "Please select a centre.")
            return redirect('manage_students')
        centre = get_object_or_404(Centre, id=centre_id)
    else:
        centre = request.user.centre

    school = get_object_or_404(School, id=school_id, centre=centre)
    uploaded_file = request.FILES['file']

    if not (uploaded_file.name.endswith('.csv') or uploaded_file.name.endswith('.xlsx')):
        messages.error(request, "Please upload a CSV or Excel file.")
        return redirect('manage_students')

    try:
        students_data = []
        errors = []
        created_count = 0

        # Parse file
        if uploaded_file.name.endswith('.csv'):
            text_file = TextIOWrapper(uploaded_file.file, encoding='utf-8')
            reader = csv.DictReader(text_file)
            students_data = list(reader)
        else:
            wb = openpyxl.load_workbook(uploaded_file)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                students_data.append(dict(zip(headers, row)))

        if not students_data:
            messages.error(request, "File is empty.")
            return redirect('manage_students')

        with transaction.atomic():
            for idx, row in enumerate(students_data, start=2):
                first_name = str(row.get('first_name', '')).strip()
                last_name = str(row.get('last_name', '') or '').strip()
                child_ID = str(row.get('child_ID', '')).strip()
                grade = str(row.get('grade', '') or '').strip()

                if not all([first_name, child_ID]):
                    errors.append(f"Row {idx}: first_name and child_ID required")
                    continue

                if Student.objects.filter(child_ID=child_ID).exists():
                    errors.append(f"Row {idx}: child_ID {child_ID} already exists")
                    continue

                if grade and grade not in dict(Student.GRADE_CHOICES):
                    errors.append(f"Row {idx}: Invalid grade '{grade}'")
                    continue

                try:
                    student = Student(
                        child_ID=child_ID,
                        name=f"{first_name} {last_name}".strip(),
                        centre=centre,
                        school=school,
                        grade=grade if grade else None
                    )
                    student.save()  # Creates user via signal

                    # Set password = child_ID
                    student.user.set_password(child_ID)
                    student.user.force_password_change = True
                    student.user.save()

                    created_count += 1
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")

        if created_count:
            messages.success(request, f"{created_count} students uploaded successfully!")
        if errors:
            msg = "\n".join(errors[:10])
            if len(errors) > 10:
                msg += f"\n... and {len(errors)-10} more"
            messages.error(request, msg)

    except Exception as e:
        messages.error(request, f"Error processing file: {str(e)}")

    return redirect('manage_students')

@login_required
def student_update(request, pk):
    """Update student — UPDATED FOR NEW MODELS"""
    if not is_authorized(request.user):
        messages.error(request, "You do not have permission.")
        return redirect('manage_students')

    student = get_object_or_404(Student, pk=pk, centre=request.user.centre)

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        child_ID = request.POST.get('child_ID', '').strip()
        school_id = request.POST.get('school', '').strip()
        grade = request.POST.get('grade', '').strip()

        if not all([first_name, child_ID, school_id]):
            messages.error(request, "All fields required.")
            return redirect('manage_students')

        if Student.objects.filter(child_ID=child_ID).exclude(pk=pk).exists():
            messages.error(request, f"Child ID {child_ID} already in use.")
            return redirect('manage_students')

        try:
            school = get_object_or_404(School, id=school_id, centre=request.user.centre)

            with transaction.atomic():
                student.name = f"{first_name} {last_name}".strip()
                old_child_ID = student.child_ID
                student.child_ID = child_ID
                student.school = school
                student.grade = grade if grade in dict(Student.GRADE_CHOICES) else None
                student.save()  # This updates user.login_id automatically

                # Update password if child_ID changed
                if old_child_ID != child_ID:
                    student.user.set_password(child_ID)
                    student.user.save()

                messages.success(request, f"Student {student.name} updated!")
            return redirect('manage_students')

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return redirect('manage_students')

    schools = School.objects.filter(centre=request.user.centre)
    context = {
        'student': student,
        'schools': schools,
        'grade_choices': Student.GRADE_CHOICES,
    }
    return render(request, 'students/student_update.html', context)

@login_required
def student_delete(request, pk):
    """Delete a student — ALREADY CORRECT"""
    if not is_authorized(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('dashboard')

    student = get_object_or_404(Student, pk=pk, centre=request.user.centre)

    if request.method == 'POST':
        student_name = student.name
        user = student.user
        with transaction.atomic():
            student.delete()        # Deletes Student
            if user:                # Deletes associated CustomUser
                user.delete()
        messages.success(request, f"Student {student_name} deleted successfully!")
        return redirect('manage_students')
    
    return redirect('manage_students')


@login_required
def download_sample_excel(request):
    """Download updated sample Excel template — NO EMAIL, LOGIN = CHILD_ID"""
    if not is_authorized(request.user):
        messages.error(request, "You do not have permission.")
        return redirect('manage_students')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    # Updated headers
    headers = ['first_name', 'last_name', 'child_ID', 'grade']
    ws.append(headers)

    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="143C50", end_color="143C50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample data
    sample_data = [
        ['John', 'Doe', '1001', 'K'],
        ['Jane', 'Smith', '1002', '1'],
        ['Bob', 'Johnson', '1003', '2'],
    ]
    for row in sample_data:
        ws.append(row)

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    # Updated Instructions
    instructions_ws = wb.create_sheet("Instructions")
    instructions = [
        ["BULK STUDENT UPLOAD - UPDATED 2025"],
        [""],
        ["IMPORTANT: Students now login using their Child ID only"],
        ["No email address is required or used"],
        [""],
        ["Column Requirements:"],
        ["first_name", "Student's first name (required)"],
        ["last_name", "Student's last name (optional)"],
        ["child_ID", "Unique student ID (required) → This is their LOGIN"],
        ["grade", "Grade level (optional): K, 1, 2, ..., 12"],
        [""],
        ["Login & Password Rules:"],
        ["• Login ID = child_ID (e.g. 1001)"],
        ["• Initial password = child_ID (e.g. 1001)"],
        ["• Students MUST change password on first login"],
        ["• No email address is created or needed"],
        [""],
        ["Upload Steps:"],
        ["1. Select Centre and School in the upload form"],
        ["2. Upload this file"],
        ["3. Students can immediately login with their child_ID"],
    ]
    for row in instructions:
        instructions_ws.append(row)

    instructions_ws.column_dimensions['A'].width = 35
    instructions_ws.column_dimensions['B'].width = 60

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_upload_template_2025.xlsx"'
    wb.save(response)
    return response
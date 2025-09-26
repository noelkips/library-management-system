# books/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.utils import timezone
import csv
import openpyxl
from io import TextIOWrapper
from .models import Book, Centre, CustomUser

def is_authorized(user):
    return user.is_superuser or user.is_staff

def handle_uploaded_file(file, user):
    header_mapping = {
        'title': 'title',
        'author': 'author',
        'category': 'category',
        'book_code': 'book_code',
        'publisher': 'publisher',
        'year_of_publication': 'year_of_publication',
        'total_copies': 'total_copies',
        'available_copies': 'available_copies',
        'centre_code': 'centre_code',
    }

    book_instances = []
    errors = []

    try:
        if file.name.lower().endswith('.csv'):
            file.seek(0)
            decoded_file = TextIOWrapper(file.file, encoding='utf-8-sig')
            reader = csv.reader(decoded_file)
            headers = next(reader, None)
            if not headers:
                raise ValueError("File is empty or invalid.")
            headers = [h.lower().strip() for h in headers]

            for row in reader:
                if not any(row):
                    continue
                data = {header_mapping.get(header): value.strip() if value else None for header, value in zip(headers, row) if header in header_mapping}
                book_instances.append(data)

        elif file.name.lower().endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
            for row in ws.iter_rows(min_row=2):
                data = {header_mapping.get(headers[i]): cell.value if cell.value else None for i, cell in enumerate(row) if i < len(headers) and headers[i] in header_mapping}
                book_instances.append(data)

        else:
            raise ValueError("Unsupported file format. Only CSV and XLSX are allowed.")

        created_count = 0
        for data in book_instances:
            centre_code = data.get('centre_code')
            centre = None
            if centre_code:
                try:
                    centre = Centre.objects.get(centre_code=centre_code)
                except Centre.DoesNotExist:
                    errors.append(f"Centre with code {centre_code} not found.")
                    continue

            if user.is_staff and not user.is_superuser and user.centre and centre != user.centre:
                errors.append("You can only add books for your own centre.")
                continue

            try:
                with transaction.atomic():
                    book = Book(
                        title=data.get('title'),
                        author=data.get('author'),
                        category=data.get('category'),
                        book_code=data.get('book_code'),
                        publisher=data.get('publisher'),
                        year_of_publication=data.get('year_of_publication'),
                        total_copies=data.get('total_copies') or 1,
                        available_copies=data.get('available_copies') or data.get('total_copies') or 1,
                        centre=centre or (user.centre if user.is_staff and not user.is_superuser else None),
                        added_by=user,
                    )
                    book.save()
                    created_count += 1
            except IntegrityError:
                errors.append(f"Book with code {data.get('book_code')} already exists in the centre.")
            except ValueError as ve:
                errors.append(str(ve))

        if created_count > 0:
            messages.success(None, f"{created_count} books imported successfully.")
        if errors:
            for error in errors:
                messages.error(None, error)

    except Exception as e:
        messages.error(None, f"Error processing file: {str(e)}")

@login_required
def book_list(request):
    if not is_authorized(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('dashboard')

    if request.user.is_superuser:
        books = Book.objects.all()
    elif request.user.is_staff:
        books = Book.objects.filter(centre=request.user.centre)
    else:
        books = Book.objects.none()

    search_query = request.GET.get('search', '')
    if search_query:
        query = (
            Q(title__icontains=search_query) |
            Q(author__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(book_code__icontains=search_query) |
            Q(publisher__icontains=search_query) |
            Q(year_of_publication__icontains=search_query) |
            Q(centre__name__icontains=search_query) |
            Q(centre__centre_code__icontains=search_query)
        )
        books = books.filter(query)

    items_per_page = int(request.GET.get('items_per_page', 10))
    paginator = Paginator(books, items_per_page)
    page_number = request.GET.get('page', 1)

    try:
        books_page = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        books_page = paginator.page(1)

    context = {
        'books': books_page,
        'paginator': paginator,
        'search_query': search_query,
        'items_per_page': items_per_page,
        'items_per_page_options': [10, 25, 50, 100],
    }
    return render(request, 'books/book_list.html', context)

@login_required
def book_add(request):
    if not is_authorized(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('dashboard')

    centres = Centre.objects.all() if request.user.is_superuser else [request.user.centre] if request.user.centre else []

    if request.method == 'POST':
        if 'file' in request.FILES:
            handle_uploaded_file(request.FILES['file'], request.user)
            return redirect('book_list')
        else:
            try:
                centre_id = request.POST.get('centre')
                centre = Centre.objects.get(id=centre_id) if centre_id else None
                if request.user.is_staff and not request.user.is_superuser and centre != request.user.centre:
                    messages.error(request, "You can only add books for your own centre.")
                    return redirect('book_add')

                book = Book(
                    title=request.POST.get('title'),
                    author=request.POST.get('author'),
                    category=request.POST.get('category'),
                    book_code=request.POST.get('book_code'),
                    publisher=request.POST.get('publisher'),
                    year_of_publication=request.POST.get('year_of_publication'),
                    total_copies=request.POST.get('total_copies') or 1,
                    available_copies=request.POST.get('available_copies') or request.POST.get('total_copies') or 1,
                    centre=centre,
                    added_by=request.user,
                )
                book.save()
                messages.success(request, "Book added successfully.")
                return redirect('book_list')
            except IntegrityError:
                messages.error(request, "Book code already exists in the centre.")
            except Exception as e:
                messages.error(request, f"Error adding book: {str(e)}")

    return render(request, 'books/book_add.html', {'centres': centres})

@login_required
def book_update(request, pk):
    if not is_authorized(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('dashboard')

    book = get_object_or_404(Book, pk=pk)
    if request.user.is_staff and not request.user.is_superuser and book.centre != request.user.centre:
        messages.error(request, "You can only update books for your own centre.")
        return redirect('book_list')

    centres = Centre.objects.all() if request.user.is_superuser else [request.user.centre] if request.user.centre else []

    if request.method == 'POST':
        try:
            centre_id = request.POST.get('centre')
            centre = Centre.objects.get(id=centre_id) if centre_id else None
            if request.user.is_staff and not request.user.is_superuser and centre != request.user.centre:
                messages.error(request, "You can only update books for your own centre.")
                return redirect('book_update', pk=pk)

            book.title = request.POST.get('title', book.title)
            book.author = request.POST.get('author', book.author)
            book.category = request.POST.get('category', book.category)
            book.book_code = request.POST.get('book_code', book.book_code)
            book.publisher = request.POST.get('publisher', book.publisher)
            book.year_of_publication = request.POST.get('year_of_publication', book.year_of_publication)
            book.total_copies = request.POST.get('total_copies', book.total_copies)
            book.available_copies = request.POST.get('available_copies', book.available_copies)
            book.centre = centre or book.centre
            book.save()
            messages.success(request, "Book updated successfully.")
            return redirect('book_list')
        except IntegrityError:
            messages.error(request, "Book code already exists in the centre.")
        except Exception as e:
            messages.error(request, f"Error updating book: {str(e)}")

    return render(request, 'books/book_update.html', {'book': book, 'centres': centres})

@login_required
def book_delete(request, pk):
    if not is_authorized(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('dashboard')

    book = get_object_or_404(Book, pk=pk)
    if request.user.is_staff and not request.user.is_superuser and book.centre != request.user.centre:
        messages.error(request, "You can only delete books for your own centre.")
        return redirect('book_list')

    if request.method == 'POST':
        book.delete()
        messages.success(request, "Book deleted successfully.")
        return redirect('book_list')

    return redirect('book_list')